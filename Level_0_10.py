import os
from openpyxl import Workbook, load_workbook

def PrintingCosts(Line):
    try:
        assert type(Line) is str

        ARR = list(Line)
        paht = os.path.dirname(os.path.abspath(__file__)) + '\Lib\ConsTon.xlsx'  # Получить директорию, где расположен скрипт + каталог с файлом xls
        WB = load_workbook(filename = paht)  # Открываем книгу
        SH = WB.active  # Переменная sheet = активному листу в книге wb 
        
        row_max = SH.max_row  # Получаем количество столбцов 
        column_max = SH.max_column  # Получаем количество строк

        sum = 0  # итоговая сумма
        for i in range(len(ARR)):  # i-ый эл-т строки
            fl = False  # флаг для отслеживания изменения суммы
            for j in range(1, row_max + 1):  # j-ая строка в файле
                if ARR[i] == SH.cell(row=j, column=1).value:
                    sum += SH.cell(row=j, column=2).value
                    fl = True
                    break
            if fl == False:
                sum += 23
        
        return sum
    
    except AssertionError:
        pass